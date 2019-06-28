# -*- coding: utf-8 -*-
"""
Created on Mon Jul 09 14:48:28 2018

@author: kasibhatlas
"""

import pandas as pd
import sys
import numpy as np
import datetime
from copy import deepcopy
import os
import calendar
import openpyxl
import openpyxl.styles.borders
import openpyxl.styles.numbers
from openpyxl.styles.fonts import Font
from openpyxl.styles import PatternFill, Border, Side, Alignment
from tabulate import tabulate


Logic = 'YTD'
month = 'June'
dateString = '27'
#Month you wanted to Run. (for running on 2nd business day of Feb 2019, Month_To_Be_Run = 1 , Year_To_Be_Run = 2019)
Month_To_Be_Run = 6
#Year You wanted to Run
Year_To_Be_Run = 2018

projectPath = r'C:\Users\nardekars\Documents\icarus-master\SRG reports\\'

todaysDirectory = projectPath + str(month) +'_'+str(dateString)

Interim_File_Path = todaysDirectory + r'\interim_files\\'

Input_File_Path = todaysDirectory + r'\input_files\\'

#Client_Path=r'C:\Users\nardekars.BSG\Documents\icarus-master\SRG reports\MF Interims Changed Logic\to Srikanth\apr_2_2019_1\Client1\\'
Cusip_Path= todaysDirectory + r'\EQ_results\\'
#Interim_File_Path=r'C:\Users\nardekars.BSG\Documents\icarus-master\SRG reports\MF Interims Changed Logic\to Srikanth\apr_2_2019\interim_files\\'

def MakeDirectory(directory):
    if not os.path.exists(directory):
        os.makedirs(directory)   

MakeDirectory(todaysDirectory)
MakeDirectory(Interim_File_Path)
MakeDirectory(Input_File_Path)        
#==============================================================================
# Sorting & Mapping related Function definations.
#==============================================================================

def SortandMap(dfCYTemp,dfPYTemp,dfPY,fyCYSameCusip1Main,jobsMapped,dfCY):
        dfCYTempSorted = dfCYTemp.sort_values('process_items',ascending=False)
        dfPYTempSorted = dfPYTemp.sort_values('process_items',ascending=False)
        dfCYTempSorted = dfCYTempSorted.reset_index()
        dfPYTempSorted = dfPYTempSorted.reset_index()
        del dfCYTempSorted['index']
        del dfPYTempSorted['index']

        if (len(dfCYTempSorted) > len(dfPYTempSorted)):
            smallerLength = len(dfPYTempSorted)
        else:
            smallerLength = len(dfCYTempSorted)
        
        dfCYTempSorted['match'] = ''
        dfCYTempSorted['py_ml_date'] = ''
        dfCYTempSorted['py_pieces'] = ''
        dfCYTempSorted['py_job'] = ''
        for index in range(0,smallerLength):
            dfCYTempSorted['match'].ix[index] = 'Y'
            dfCYTempSorted['py_ml_date'].ix[index] = dfPYTempSorted.ix[index]['ml_date']
            dfCYTempSorted['py_pieces'].ix[index] = dfPYTempSorted.ix[index]['process_items']
            dfCYTempSorted['py_job'].ix[index] = dfPYTempSorted.ix[index]['job_number']
            jobsMapped.append(dfPYTempSorted.ix[index]['job_number'])
            jobsMapped.append(dfCYTempSorted.ix[index]['job_number'])
        
        dfPY = dfPY[~dfPY['job_number'].isin(jobsMapped)]
        dfCY = dfCY[~dfCY['job_number'].isin(jobsMapped)]
        
        fyCYSameCusip1Main = pd.concat([fyCYSameCusip1Main,dfCYTempSorted])   
        
        return fyCYSameCusip1Main,dfPY,jobsMapped,dfCY
        
'''

'''

def Precedence(dfCY,dfPY):
    jobsMapped = []
    fyCYSameCusip1Main = pd.DataFrame()
    
    if (len(dfCY[dfCY['jobtype']=='P'])>0) & (len(dfPY[dfPY['jobtype']=='Z'])>0):
        dfCYTemp = dfCY[dfCY['jobtype']=='P']
        dfPYTemp = dfPY[dfPY['jobtype']=='Z']
        
        fyCYSameCusip1Main,dfPY,jobsMapped,dfCY = SortandMap(dfCYTemp,dfPYTemp,dfPY,fyCYSameCusip1Main,jobsMapped,dfCY)     
    
    if ((len(dfCY[dfCY['jobtype']=='P'])>0) & (len(dfPY[dfPY['jobtype']=='S'])>0)):
        dfCYTemp = dfCY[dfCY['jobtype']=='P']
        dfPYTemp = dfPY[dfPY['jobtype']=='S']

        fyCYSameCusip1Main,dfPY,jobsMapped,dfCY = SortandMap(dfCYTemp,dfPYTemp,dfPY,fyCYSameCusip1Main,jobsMapped,dfCY) 
                       
    
    if (len(dfCY[dfCY['jobtype']=='Z'])>0) & (len(dfPY[dfPY['jobtype']=='S'])>0):
        dfCYTemp = dfCY[dfCY['jobtype']=='Z']
        dfPYTemp = dfPY[dfPY['jobtype']=='S']

        fyCYSameCusip1Main,dfPY,jobsMapped,dfCY = SortandMap(dfCYTemp,dfPYTemp,dfPY,fyCYSameCusip1Main,jobsMapped,dfCY)  
                      
    
    if (len(dfCY[dfCY['jobtype']=='S'])>0) & (len(dfPY[dfPY['jobtype']=='Z'])>0):
        dfCYTemp = dfCY[dfCY['jobtype']=='S']
        dfPYTemp = dfPY[dfPY['jobtype']=='Z']
        
        fyCYSameCusip1Main,dfPY,jobsMapped,dfCY = SortandMap(dfCYTemp,dfPYTemp,dfPY,fyCYSameCusip1Main,jobsMapped,dfCY) 
                
    
    if (len(dfCY[dfCY['jobtype']=='S'])>0) & (len(dfPY[dfPY['jobtype']=='P'])>0):
        dfCYTemp = dfCY[dfCY['jobtype']=='S']
        dfPYTemp = dfPY[dfPY['jobtype']=='P']

        fyCYSameCusip1Main,dfPY,jobsMapped,dfCY = SortandMap(dfCYTemp,dfPYTemp,dfPY,fyCYSameCusip1Main,jobsMapped,dfCY)   
                
    
    if (len(dfCY[dfCY['jobtype']=='Z'])>0) & (len(dfPY[dfPY['jobtype']=='P'])>0):
        dfCYTemp = dfCY[dfCY['jobtype']=='Z']
        dfPYTemp = dfPY[dfPY['jobtype']=='P']


        fyCYSameCusip1Main,dfPY,jobsMapped,dfCY = SortandMap(dfCYTemp,dfPYTemp,dfPY,fyCYSameCusip1Main,jobsMapped,dfCY) 
                
    return fyCYSameCusip1Main
    
    
#==============================================================================
# Read the File
#==============================================================================

NowDate = datetime.datetime.now()

Month=month
#Month_To_Be_Run = 3
#Year You wanted to Run
Year_To_Be_Run = 2018

NowDate = datetime.datetime.now()
Date_To_Run = NowDate.replace(day = 1,month=Month_To_Be_Run,year=Year_To_Be_Run)
currentMonth = datetime.datetime.now().month
currentYear= datetime.datetime.now().year

NowDate= NowDate.strftime("%Y-%m-%d")         
NowDateTime = datetime.datetime.now()
NowDateTime=NowDateTime.strftime("%Y-%m-%d %H-%M-%S")

Merged_file=pd.DataFrame()

TodaysDirecotry_Cusip = Cusip_Path+NowDate
MakeDirectory(TodaysDirecotry_Cusip)
#TodaysDirecotry_Client = Client_Path+NowDate
#MakeDirectory(TodaysDirecotry_Client)

#=============================================================================   
#====================             Logic to pick QTD/MTD/YTD Dates ===========================
#=============================================================================
#MTD Logic 

First_Day_MTD_CY = Date_To_Run.replace(day = 1,month=Month_To_Be_Run,year=Year_To_Be_Run)
Last_Day_MTD_CY = First_Day_MTD_CY.replace(day = calendar.monthrange(First_Day_MTD_CY.year, First_Day_MTD_CY.month)[1])
Last_Day_MTD_CY= Last_Day_MTD_CY.strftime("%Y-%m-%d") 
First_Day_MTD_CY= First_Day_MTD_CY.strftime("%Y-%m-%d") 

First_Day_MTD_PY = Date_To_Run.replace(day = 1,month=Month_To_Be_Run,year=Year_To_Be_Run-1)
Last_Day_MTD_PY = First_Day_MTD_PY.replace(day = calendar.monthrange(First_Day_MTD_PY.year, First_Day_MTD_PY.month)[1])
Last_Day_MTD_PY= Last_Day_MTD_PY.strftime("%Y-%m-%d") 
First_Day_MTD_PY= First_Day_MTD_PY.strftime("%Y-%m-%d") 

#print (tabulate([["MTD",First_Day_MTD_CY, Last_Day_MTD_CY,First_Day_MTD_PY,Last_Day_MTD_PY]], headers=['MTD/QTD/YTD?','CY SD','CY ED','PY SD','PY ED'], tablefmt='orgtbl'))
#YTD Logic 
if Month_To_Be_Run in [7,8,9,10,11,12]:
    Year_for_YTD_Firstday=Year_To_Be_Run
    Year_for_YTD_Lastday=Year_To_Be_Run
elif Month_To_Be_Run in [1,2,3,4,5,6]:
    Year_for_YTD_Firstday=Year_To_Be_Run-1
    Year_for_YTD_Lastday=Year_To_Be_Run
    
First_Day_YTD_CY = Date_To_Run.replace(day = 1,month=7,year=Year_for_YTD_Firstday)
#Last_Day_YTD_CY = First_Day_YTD_CY.replace(day = calendar.monthrange(Year_for_YTD_Lastday, Month_To_Be_Run)[1],month=Month_To_Be_Run,year=Year_for_YTD_Lastday)

First_Day_YTD_PY = Date_To_Run.replace(day = 1,month=7,year=Year_for_YTD_Firstday-1)
#Last_Day_YTD_PY = First_Day_YTD_PY.replace(day = calendar.monthrange(Year_for_YTD_Lastday-1, Month_To_Be_Run)[1],month=Month_To_Be_Run,year=Year_for_YTD_Lastday-1)

First_Day_YTD_CY= First_Day_YTD_CY.strftime("%Y-%m-%d") 
First_Day_YTD_PY= First_Day_YTD_PY.strftime("%Y-%m-%d") 
Last_Day_YTD_CY= Last_Day_MTD_CY 
Last_Day_YTD_PY= Last_Day_MTD_PY 
Last_Day_Whole_YTD_CY = Date_To_Run.replace(day = 30,month=6,year=Year_for_YTD_Firstday+1)
Last_Day_Whole_YTD_CY= Last_Day_Whole_YTD_CY.strftime("%Y-%m-%d") 
Last_Day_Whole_YTD_PY = Date_To_Run.replace(day = 30,month=6,year=Year_for_YTD_Firstday)
Last_Day_Whole_YTD_PY= Last_Day_Whole_YTD_PY.strftime("%Y-%m-%d") 

#QTD Logic 

Q1=[7,8,9]
Q2=[10,11,12]
Q3=[1,2,3]
Q4=[4,5,6]

if Month_To_Be_Run in Q1:
    QTD=Q1
elif Month_To_Be_Run in Q2:
    QTD=Q2
elif Month_To_Be_Run in Q3:
    QTD=Q3
elif Month_To_Be_Run in Q4:
    QTD=Q4 
    
if Month_To_Be_Run in (7,8,9,10,11,12):    
    First_Day_QTD_CY = Date_To_Run.replace(day = 1,month=QTD[0],year=Year_for_YTD_Firstday) 
    #Last_Day_QTD_CY = First_Day_MTD_CY.replace(day = calendar.monthrange(First_Day_MTD_CY.year, First_Day_MTD_CY.month)[1])
    First_Day_QTD_PY = Date_To_Run.replace(day = 1,month=QTD[0],year=Year_for_YTD_Firstday-1) 
    #Last_Day_QTD_PY = First_Day_YTD_PY.replace(day = calendar.monthrange(Year_for_YTD_Lastday-1, Month_To_Be_Run)[1],month=Month_To_Be_Run,year=Year_for_YTD_Lastday-1)
    First_Day_QTD_CY= First_Day_QTD_CY.strftime("%Y-%m-%d") 
    First_Day_QTD_PY= First_Day_QTD_PY.strftime("%Y-%m-%d") 
    Last_Day_QTD_CY= Last_Day_MTD_CY
    Last_Day_QTD_PY= Last_Day_MTD_PY 
elif Month_To_Be_Run in (1,2,3,4,5,6):
    First_Day_QTD_CY = Date_To_Run.replace(day = 1,month=QTD[0],year=Year_for_YTD_Firstday+1) 
    #Last_Day_QTD_CY = First_Day_MTD_CY.replace(day = calendar.monthrange(First_Day_MTD_CY.year, First_Day_MTD_CY.month)[1])
    First_Day_QTD_PY = Date_To_Run.replace(day = 1,month=QTD[0],year=Year_for_YTD_Firstday) 
    #Last_Day_QTD_PY = First_Day_YTD_PY.replace(day = calendar.monthrange(Year_for_YTD_Lastday-1, Month_To_Be_Run)[1],month=Month_To_Be_Run,year=Year_for_YTD_Lastday-1)
    First_Day_QTD_CY= First_Day_QTD_CY.strftime("%Y-%m-%d") 
    First_Day_QTD_PY= First_Day_QTD_PY.strftime("%Y-%m-%d") 
    Last_Day_QTD_CY= Last_Day_MTD_CY
    Last_Day_QTD_PY= Last_Day_MTD_PY  

print ("For the given month "+ str(Month_To_Be_Run)+" and Year "+str(Year_To_Be_Run)+" The below are the SD/ED for QTD/MTD/YTD")
print (tabulate([["MTD",First_Day_MTD_CY, Last_Day_MTD_CY,First_Day_MTD_PY,Last_Day_MTD_PY],["QTD",First_Day_QTD_CY, Last_Day_QTD_CY,First_Day_QTD_PY,Last_Day_QTD_PY],["YTD",First_Day_YTD_CY, Last_Day_YTD_CY,First_Day_YTD_PY,Last_Day_YTD_PY]], headers=['MTD/QTD/YTD?','CY SD','CY ED','PY SD','PY ED'], tablefmt='orgtbl'))

#==============================================================================
# CUSIP LEVEL
#==============================================================================
#grouped1 = pd.read_csv(r'C:\Data Science\SRG\Input files from GP\EQ\June 2018\Equity_Query_without_client_Result_fy17.csv',encoding='utf-8')
#grouped2 = pd.read_csv(r'C:\Data Science\SRG\Input files from GP\EQ\June 2018\Equity_Query_without_client_Result_fy18.csv',encoding='utf-8')
#grouped=pd.concat([grouped1,grouped2])

grouped = pd.read_csv(Input_File_Path + r'\EQ_Query_Result_without_client_until_'+str(month)+str(dateString)+'.csv',encoding='utf-8')

#################################
grouped = pd.read_csv(r'C:\Users\nardekars\Documents\icarus-master\SRG reports\June_27\input_files\FY_18_EQ_Query_Result_without_client_until_July01.csv')
#################################
grouped.head()

################################# 
#grouped = grouped[grouped['service']=='SH']
#################################

grouped.shape

grouped.columns=['job_number','jobtype','cusip6','issuer_name','exch_name','status','service','ml_date','record_date','jbrc_location','process_items','mailed_items','evs_consol_items','wrap_items','householding_items','nomail_items','email_items','items']

grouped ['job_number']=grouped ['job_number'].map(str).apply(lambda x: x.zfill(6))
grouped ['cusip6']=grouped ['cusip6'].map(str).apply(lambda x: x.zfill(6))
grouped['matching_column']=grouped['job_number'].map(str)+grouped['cusip6'].map(str)

grouped_pivoted = pd.pivot_table(grouped,index="matching_column",values=["process_items"],aggfunc=[np.sum],fill_value=0).reset_index()


#grouped.to_excel(r'C:\Data Science\SRG\Input files from GP\EQ\June 2018\Equity_Query_without_client_Result.xlsx')
#==============================================================================
# populate the F/C , Fiscal_Yr columns
#==============================================================================
grouped['ml_date'] =  pd.to_datetime(grouped['ml_date'], format='%Y-%m-%d')
grouped['F/C'] = ''
grouped['F/C'][grouped['exch_name'].astype(str).map(str.strip)=='FUND'] = 'F'
grouped['F/C'][grouped['exch_name'].astype(str).map(str.strip)!='FUND'] = 'C'

grouped['fiscal_yr'] = ''
grouped['fiscal_yr'][(grouped['ml_date']>=First_Day_YTD_CY) & (grouped['ml_date']<=Last_Day_Whole_YTD_CY)] = '2019'
grouped['fiscal_yr'][(grouped['ml_date']>=First_Day_YTD_PY) & (grouped['ml_date']<=Last_Day_Whole_YTD_PY)] = '2018' 

#################################
grouped['fiscal_yr'] = ''
grouped['fiscal_yr'][(grouped['ml_date']>=First_Day_YTD_CY) & (grouped['ml_date']<=Last_Day_Whole_YTD_CY)] = '2018'
grouped['fiscal_yr'][(grouped['ml_date']>=First_Day_YTD_PY) & (grouped['ml_date']<=Last_Day_Whole_YTD_PY)] = '2017' 
#################################
print('After populating the FY & F/C the cusip level Data '+ str(len(grouped)))


#removing certain jobs
jobsToExcludeData = pd.read_excel(r'C:\Users\nardekars\Documents\icarus-master\SRG reports\May_02\interim_files\Jobs to be excluded.xlsx')

CYJobsToExclude = jobsToExcludeData['FY19 Jobs'].tolist()
PYJobsToExclude = jobsToExcludeData['FY18 Jobs'].tolist()

groupedCY = grouped[grouped['fiscal_yr'] == '2019']
groupedPY = grouped[grouped['fiscal_yr'] == '2018']

#################################
groupedCY = grouped[grouped['fiscal_yr'] == '2018']
groupedPY = grouped[grouped['fiscal_yr'] == '2017']
#################################

len(groupedCY) + len(groupedPY) == len(grouped)

groupedCY = groupedCY[~groupedCY['job_number'].isin(CYJobsToExclude)]
groupedPY = groupedPY[~groupedPY['job_number'].isin(PYJobsToExclude)]

pd.concat([groupedCY,groupedPY]).shape

grouped = pd.concat([groupedCY,groupedPY])

print('After removing certain job numbers: '+ str(grouped.shape))
#==============================================================================
# Filter with exchange/Service/jbrc loation
#==============================================================================
grouped = grouped[grouped['F/C']=='C']


grouped = grouped[(grouped['service'].astype(str).map(str.strip)=='BN')|(grouped['service'].astype(str).map(str.strip)=='BD')]

#################################
groupedBNBD = grouped[(grouped['service'].astype(str).map(str.strip)=='BN')|(grouped['service'].astype(str).map(str.strip)=='BD')]
groupedSH = grouped[grouped['service'].astype(str).map(str.strip)=='SH']

groupedBNBD = groupedBNBD.set_index('issuer_name').to_dict()
groupedSH['cusip6_BNBD'] = groupedSH['issuer_name'].map(groupedBNBD['cusip6'])

notMatched = groupedSH[groupedSH['cusip6_BNBD'].astype(str).map(str.strip)=='nan']
matched = groupedSH[groupedSH['cusip6_BNBD'].astype(str).map(str.strip)!='nan']

notMatched.to_csv(r'C:\Users\nardekars\Documents\icarus-master\SRG reports\June_27\EQ_results\2019-06-27\Report\SH_notMatched_issuers.csv',index=False)

notMatched = pd.read_excel(r'C:\Users\nardekars\Documents\icarus-master\SRG reports\June_27\interim_files\SH Not Matched.xlsx')

groupedSH1 = pd.concat([matched,notMatched])
groupedSH1.shape
grouped = groupedSH1
grouped = grouped.rename(columns={'cusip6':'cusip6_SH'})
grouped = grouped.rename(columns={'cusip6_BNBD':'cusip6'})
#################################


grouped = grouped[(grouped['jbrc_location'].astype(str).map(str.strip)!='OMNI')&(grouped['jbrc_location'].astype(str).map(str.strip)!='ONMI')&(grouped['jbrc_location'].astype(str).map(str.strip)!='TEST')]

grouped= grouped[grouped['cusip6']!='QAAQAA']

print('After Filtering the Funds/Service/location Cusip level Data '+ str(len(grouped)))






#==============================================================================
# Load the client level files
#==============================================================================

Equity_Query_client=pd.read_csv(Input_File_Path + r'\EQ_Query_Result_with_client_until_'+str(month)+str(dateString)+'.csv',names=['job_number','jobtype','cusip6','client_no','client_name','issuer_name','exch_name','status','service','ml_date','record_date','jbrc_location','process_items','mailed_items','evs_consol_items','wrap_items','householding_items','nomail_items','email_items','items'])

#################################
Equity_Query_client=pd.read_csv(r'C:\Users\nardekars\Documents\icarus-master\SRG reports\June_27\input_files\FY_18_EQ_Query_Result_with_client_until_July01.csv',names=['job_number','jobtype','cusip6','client_no','client_name','issuer_name','exch_name','status','service','ml_date','record_date','jbrc_location','process_items','mailed_items','evs_consol_items','wrap_items','householding_items','nomail_items','email_items','items'])
#################################

Equity_Query_client.shape

################################# 
Equity_Query_client = Equity_Query_client[Equity_Query_client['service']=='SH']
#################################

Equity_Query_client['job_number']=Equity_Query_client ['job_number'].map(str).apply(lambda x: x.zfill(6))
Equity_Query_client['cusip6']=Equity_Query_client ['cusip6'].map(str).apply(lambda x: x.zfill(6))
Equity_Query_client['matching_column']=Equity_Query_client['job_number'].map(str)+Equity_Query_client['cusip6'].map(str)

Equity_Query_client['fiscal_yr'] = ''
Equity_Query_client['fiscal_yr'][(Equity_Query_client['ml_date']>=First_Day_YTD_CY) & (Equity_Query_client['ml_date']<=Last_Day_Whole_YTD_CY)] = '2019'
Equity_Query_client['fiscal_yr'][(Equity_Query_client['ml_date']>=First_Day_YTD_PY) & (Equity_Query_client['ml_date']<=Last_Day_Whole_YTD_PY)] = '2018' 

Equity_Query_clientCY = Equity_Query_client[Equity_Query_client['fiscal_yr'] == '2019']
Equity_Query_clientPY = Equity_Query_client[Equity_Query_client['fiscal_yr'] == '2018']

#################################
Equity_Query_client['fiscal_yr'] = ''
Equity_Query_client['fiscal_yr'][(Equity_Query_client['ml_date']>=First_Day_YTD_CY) & (Equity_Query_client['ml_date']<=Last_Day_Whole_YTD_CY)] = '2018'
Equity_Query_client['fiscal_yr'][(Equity_Query_client['ml_date']>=First_Day_YTD_PY) & (Equity_Query_client['ml_date']<=Last_Day_Whole_YTD_PY)] = '2017' 

Equity_Query_clientCY = Equity_Query_client[Equity_Query_client['fiscal_yr'] == '2018']
Equity_Query_clientPY = Equity_Query_client[Equity_Query_client['fiscal_yr'] == '2017']
#################################


len(Equity_Query_clientCY) + len(Equity_Query_clientPY) == len(Equity_Query_client)

Equity_Query_clientCY = Equity_Query_clientCY[~Equity_Query_clientCY['job_number'].isin(CYJobsToExclude)]
Equity_Query_clientPY = Equity_Query_clientPY[~Equity_Query_clientPY['job_number'].isin(PYJobsToExclude)]

#pd.concat([Equity_Query_clientCY,Equity_Query_clientPY]).shape

Equity_Query_client = pd.concat([Equity_Query_clientCY,Equity_Query_clientPY])

print('After removing certain job numbers (in with client): '+ str(Equity_Query_client.shape))

#################################
import copy
Equity_Query_client1 = copy.deepcopy(Equity_Query_client)
grouped1 = grouped.set_index('cusip6').to_dict()
Equity_Query_client1['issuer_name'] = Equity_Query_client['cusip6'].map(grouped1['issuer_name'])
Equity_Query_client = copy.deepcopy(Equity_Query_client1)
#################################

Equity_Query_client_pivoted = pd.pivot_table(Equity_Query_client,index="matching_column",values=["process_items"],aggfunc=[np.sum],fill_value=0).reset_index()

Equity_Query_client_pivoted['process_items'] = Equity_Query_client_pivoted['sum']['process_items']
del Equity_Query_client_pivoted['sum']

grouped_pivoted['process_items'] = grouped_pivoted['sum']['process_items']
del grouped_pivoted['sum']

#pd.merge(Equity_Query_client_pivoted,grouped_pivoted,on='matching_column').shape

merged = pd.merge(Equity_Query_client_pivoted,grouped_pivoted,on='matching_column',how='left')

#merged['process_items_x'] = merged['process_items_x'].map(float)

differenceList = merged[merged['process_items_x'].map(int) != merged['process_items_y'].map(float).map(int)]['matching_column'].str[0:6].tolist()

for x in set(differenceList):
    print("'"+str(x)+"'",end=',')
    
if len(differenceList) > 0:
    #breaking
    
    print('breaking')
    
    sys.exit()
    
        
    
    #blList=['P86251','S64271']
    Equity_Query_client=Equity_Query_client[~(Equity_Query_client['job_number'].isin(differenceList))]
    
    #updated is the differences file (and to be added to with_client)
    updated = pd.read_csv(r'C:\Users\nardekars\Documents\icarus-master\SRG reports\June_27\input_files\FY_18_EQ_Query_Result_with_client_until_July01_difference.csv')
    #Equity_Query_client1=pd.read_csv(r'C:\Data Science\SRG\Input files from GP\EQ\June 2018\EQ_Query_Result_with_client_fy17.csv')
    #Equity_Query_client2=pd.read_csv(r'C:\Data Science\SRG\Input files from GP\EQ\June 2018\EQ_Query_Result_with_client_fy18.csv')
    #Equity_Query_client=pd.concat([Equity_Query_client1,Equity_Query_client2])
    #Equity_Query_client1.columns= ['job_number','jobtype', 'cusip6','client_no', 'client_name','issuer_name', 'exch_name', 'status', 'service', 'ml_date', 'record_date','jbrc_location','process_items', 'mailed_items', 'evs_consol_items', 'wrap_items', 'householding_items', 'nomail_items', 'email_items', 'items'] 
    #Equity_Query_client2.columns= ['job_number','jobtype', 'cusip6','client_no', 'client_name','issuer_name', 'exch_name', 'status', 'service', 'ml_date', 'record_date','jbrc_location','process_items', 'mailed_items', 'evs_consol_items', 'wrap_items', 'householding_items', 'nomail_items', 'email_items', 'items'] 
    updated.columns= ['job_number','jobtype', 'cusip6','client_no', 'client_name','issuer_name', 'exch_name', 'status', 'service', 'ml_date', 'record_date','jbrc_location','process_items', 'mailed_items', 'evs_consol_items', 'wrap_items', 'householding_items', 'nomail_items', 'email_items', 'items'] 
    
    
    client_grouped=pd.concat([updated,Equity_Query_client])
else:
    client_grouped=Equity_Query_client

######################
Equity_Query_client_pivoted = pd.pivot_table(Equity_Query_client,index="matching_column",values=["process_items"],aggfunc=[np.sum],fill_value=0).reset_index()

Equity_Query_client_pivoted['process_items'] = Equity_Query_client_pivoted['sum']['process_items']
del Equity_Query_client_pivoted['sum']

#pd.merge(Equity_Query_client_pivoted,grouped_pivoted,on='matching_column').shape

merged = pd.merge(Equity_Query_client_pivoted,grouped_pivoted,on='matching_column',how='left')

differenceList = merged[merged['process_items_x'].map(int) != merged['process_items_y'].map(float).map(int)]['matching_column'].str[0:6].tolist()

client_grouped = client_grouped[client_grouped['service']=='SH']
######################


client_grouped.shape

client_grouped.to_csv(str(Interim_File_Path)+str(month)+'_2019_with_client_updated_SH.csv',index=False)




#client_grouped = pd.read_csv(r'C:\Users\nardekars.BSG\Documents\icarus-master\SRG reports\MF Interims Changed Logic\to Srikanth\Mar_2019_with_client_updated.csv')

#==============================================================================
# populate the F/C , Fiscal_Yr columns
#==============================================================================

client_grouped['ml_date'] =  pd.to_datetime(client_grouped['ml_date'], format='%Y-%m-%d')
client_grouped['F/C'] = ''
client_grouped['F/C'][client_grouped['exch_name'].astype(str).map(str.strip)=='FUND'] = 'F'
client_grouped['F/C'][client_grouped['exch_name'].astype(str).map(str.strip)!='FUND'] = 'C'
#client_grouped['fiscal_yr'] = ''
#client_grouped['fiscal_yr'][(client_grouped['ml_date']>=First_Day_YTD_CY) & (client_grouped['ml_date']<=Last_Day_Whole_YTD_CY)] = '2019'
#client_grouped['fiscal_yr'][(client_grouped['ml_date']>=First_Day_YTD_PY) & (client_grouped['ml_date']<=Last_Day_Whole_YTD_PY)] = '2018' 

print('After populating the FY & F/C the client level Data '+ str(len(client_grouped)))

#==============================================================================
# Filter with exchange/Service/jbrc loation
#==============================================================================
client_grouped = client_grouped[client_grouped['F/C']=='C']

client_grouped = client_grouped[(client_grouped['service'].astype(str).map(str.strip)=='BN')|(client_grouped['service'].astype(str).map(str.strip)=='BD')]

client_grouped = client_grouped[(client_grouped['jbrc_location'].astype(str).map(str.strip)!='OMNI')&(client_grouped['jbrc_location'].astype(str).map(str.strip)!='ONMI')&(client_grouped['jbrc_location'].astype(str).map(str.strip)!='TEST')]
client_grouped= client_grouped[client_grouped['cusip6']!='QAAQAA']

print('After Filtering the Funds/Service/location client level Data '+ str(len(client_grouped)))

#==============================================================================
# Add process items updated columns
#==============================================================================

#client_grouped['process_items'].sum()

client_grouped['process_items_updated'] = ''
client_grouped['process_items_updated'] = client_grouped['process_items'].map(int)

if Logic == 'YTD':
    client_grouped['process_items_updated'][client_grouped['client_no']=='3VM'] = 0

client_grouped['process_items_updated'][client_grouped['client_no']=='11R'] = 0
client_grouped['process_items_updated'][client_grouped['client_no']=='4PG'] = 0
client_grouped['process_items_updated'][client_grouped['client_no']=='161'] = client_grouped[client_grouped['client_no']=='161']['process_items'].map(int) - client_grouped[client_grouped['client_no']=='161']['nomail_items'].map(int)


#handling error
if len(client_grouped[client_grouped['process_items'].astype(str).map(str.strip)!='nan']) != len(client_grouped):
    sys.exit()
else:
    print('correct')
#client_grouped = client_grouped[client_grouped['process_items'].astype(str).map(str.strip)!='nan']

client_grouped=client_grouped.reset_index()
del client_grouped['index']

print('After updating the process items '+ str(len(client_grouped)))



CY = '2018'
PY = '2017'

def MFSRGLogic(First_Day_PY,Last_Day_PY,First_Day_CY,Last_Day_CY):
    
    print('Analysis initiated for '+Logic)
    #==============================================================================
    # Split the data into FYs
    #==============================================================================
    
    fyPY = grouped[(grouped['ml_date']>=First_Day_PY) & (grouped['ml_date']<=Last_Day_PY)] 
    fyCY = grouped[(grouped['ml_date']>=First_Day_CY) & (grouped['ml_date']<=Last_Day_CY)]
    fyCY = fyCY.reset_index()
    del fyCY['index']
    fyPY = fyPY.reset_index()
    del fyPY['index']

    #==============================================================================
    # Split the data into FYs for client level data
    #==============================================================================
    FY_PY = client_grouped[(client_grouped['ml_date']>=First_Day_PY) & (client_grouped['ml_date']<=Last_Day_PY)] 
    FY_CY = client_grouped[(client_grouped['ml_date']>=First_Day_CY) & (client_grouped['ml_date']<=Last_Day_CY)] 

    FY_PY = FY_PY.reset_index()
    del FY_PY['index']
    FY_CY = FY_CY.reset_index()
    del FY_CY['index'] 
    
    #==============================================================================
    #Sorting & Mapping
    #==============================================================================
    fyCY['match'] = ''
    fyCY['py_ml_date'] = ''
    fyCY['py_pieces'] = ''
    fyCY['py_job'] = ''
    fyPY['mapped'] = ''
    
    for x in range(0,3):
        fyCY = fyCY.reset_index()
        del fyCY['index']
    fyCY = fyCY.reset_index()
    fyCY = fyCY.rename(columns={'index':'original_index'})
#    temp = fyPY
#    fyPY = temp
    
    for x in range(0,3):
        fyPY = fyPY.reset_index()
        del fyPY['index']
    fyPY = fyPY.reset_index()
    
    fyPY = fyPY.rename(columns={'index':'original_index'})
    
    groupCYByCusips = fyCY.groupby(u'cusip6').groups
#    main_df = pd.DataFrame()
    grp_counter = 0

    '''
    a.	P job to P job, S job to S job, Z job to Z job
    b.	P job to Z job
    c.	P job to S job
    d.	Z job to S job
    e.	S job to Z job
    f.	S job to P job
    g.	Z job to P job
    '''
    ########
#    original_index = 100000000000
    mapped = pd.DataFrame()
    fyCYSameCusip1Main = pd.DataFrame()
    fyPYSameCusip1Main = pd.DataFrame()
    grp_counter = 0
    
    totalLengthCusips = len(groupCYByCusips.items())
    for cusip,cusipIndices in groupCYByCusips.items():
        print (grp_counter,totalLengthCusips)
        fyCYSameCusip = fyCY.ix[cusipIndices]
        fyPYSameCusip = fyPY[fyPY['cusip6'] == cusip]
        
        fyCYSameCusip = fyCYSameCusip.reset_index()
        del fyCYSameCusip['index']
        fyPYSameCusip = fyPYSameCusip.reset_index()
        del fyPYSameCusip['index']
            
        intersectedJobs = set(fyCYSameCusip['jobtype'].tolist()).intersection(set(fyPYSameCusip['jobtype'].tolist()))
        '''if there are same jobtypes on either side, sort both sides and match all '''
        
        if len(intersectedJobs) > 0:
            for jobtype in intersectedJobs:
                fyCYSameCusip1 = fyCYSameCusip[fyCYSameCusip['jobtype'] == jobtype]
                fyPYSameCusip1 = fyPYSameCusip[fyPYSameCusip['jobtype'] == jobtype]
                fyCYSameCusip1 = fyCYSameCusip1.sort_values('process_items',ascending=False)
                fyPYSameCusip1 = fyPYSameCusip1.sort_values('process_items',ascending=False)
                
                fyCYSameCusip1 = fyCYSameCusip1.reset_index()
                del fyCYSameCusip1['index']
                fyPYSameCusip1 = fyPYSameCusip1.reset_index()
                del fyPYSameCusip1['index']
                
                if (len(fyCYSameCusip1) > len(fyPYSameCusip1)):
                    smallerLength = len(fyPYSameCusip1)
                else:
                    smallerLength = len(fyCYSameCusip1)
                    
                for index in range(0,smallerLength):
                    fyCYSameCusip1['match'].ix[index] = 'Y'
                    fyCYSameCusip1['py_ml_date'].ix[index] = fyPYSameCusip1.ix[index]['ml_date']
                    fyCYSameCusip1['py_pieces'].ix[index] = fyPYSameCusip1.ix[index]['process_items']
                    fyCYSameCusip1['py_job'].ix[index] = fyPYSameCusip1.ix[index]['job_number']
                    
                    fyPYSameCusip1['mapped'].ix[index] = 'Mapped'
                
                fyCYSameCusip1 = fyCYSameCusip1[fyCYSameCusip1['match']=='Y']
                fyCYSameCusip1Main = pd.concat([fyCYSameCusip1Main,fyCYSameCusip1])
                fyPYSameCusip1Main = pd.concat([fyPYSameCusip1Main,fyPYSameCusip1])
                mapped = pd.concat([mapped,fyCYSameCusip1Main])
        else:
            fyPYSameCusip1 = deepcopy(fyPYSameCusip)
            
        '''take the rest if some are mapped else take everything (in current year)'''
        if len(fyCYSameCusip1Main) > 0:
            fyCYSameCusipRest = fyCYSameCusip[~(fyCYSameCusip['original_index'].isin(fyCYSameCusip1Main['original_index'].tolist()))]
        else:
            fyCYSameCusipRest = deepcopy(fyCYSameCusip)
        
        '''take the rest if some are mapped else take everything (in previous year)'''
        
        if len(fyPYSameCusip1Main) > 0:
            fyPYSameCusipRest = fyPYSameCusip[~(fyPYSameCusip['job_number'].isin(fyPYSameCusip1Main[fyPYSameCusip1Main['mapped']=='Mapped']['job_number']))]
        else:
             fyPYSameCusipRest = deepcopy(fyPYSameCusip)
        #########################
        
        '''if both the sides have some rows'''
        if (len(fyCYSameCusipRest)>0) & (len(fyPYSameCusipRest)>0):
            fyCYSameCusip1Main = Precedence(fyCYSameCusipRest,fyPYSameCusipRest)
        mapped = pd.concat([mapped,fyCYSameCusip1Main])
        
        grp_counter+= 1
        
    mapped.shape
    
    fyCY['merge_column']=fyCY['job_number']+fyCY['process_items'].map(str)
    mapped['merge_column']=mapped['job_number']+mapped['process_items'].map(str)
    
    
    mappedRest = fyCY[~(fyCY['merge_column'].isin(mapped['merge_column']))]
    
    mapped = pd.concat([mapped,mappedRest])
    mapped1 = mapped.drop_duplicates()
    
    mapped1['match'][mapped1['match']!='Y'] = 'N'
    
    mapped1 = mapped1[[ u'job_number', u'jobtype', u'cusip6', u'issuer_name', u'exch_name', u'status', u'service', u'ml_date', u'record_date', u'jbrc_location', u'process_items', u'mailed_items', u'evs_consol_items', u'wrap_items', u'householding_items', u'nomail_items', u'email_items', u'items', u'F/C', u'match', u'py_ml_date', u'py_pieces', u'py_job']]

    Cusip_Result=deepcopy(mapped1)
    print ("Cusip Level Analysis is successful, Now using the Data and performing the client level Analysis for "+Logic)
   
    #==============================================================================
    #Sorting & Mapping for client level data
    #==============================================================================
    Final_Mapped=deepcopy(FY_CY)
    Final_Mapped['process_items_old_client']=Final_Mapped['process_items']
    del Final_Mapped['process_items']
    Final_Mapped['process_items']=Final_Mapped['process_items_updated']
    del Final_Mapped['process_items_updated']
    Old_result=deepcopy(mapped1)
    Old_result=Old_result[['job_number','process_items','match','py_pieces','py_job']]
    Old_result_dict=Old_result.set_index('job_number').to_dict()
    Final_Mapped['py_job'] = Final_Mapped['job_number'].map(Old_result_dict['py_job'])
    Final_Mapped['match']=''
    Final_Mapped['match'] = Final_Mapped['job_number'].map(Old_result_dict['match'])
    Final_Mapped['py_pieces_old']=Final_Mapped['job_number'].map(Old_result_dict['py_pieces'])
    Final_Mapped['Cy_pieces_old']=Final_Mapped['job_number'].map(Old_result_dict['process_items'])
    
    Final_Mapped=Final_Mapped.reset_index()
    del Final_Mapped['index']

    FY_PY['match']=''
    FY_PY = FY_PY.rename(columns={'job_number':'py_job'})
    Final_Mapped_dict_PYjob=Final_Mapped.set_index('py_job').to_dict()
#    fy2017_dict_PYjob=FY_PY.set_index('py_job').to_dict()
    FY_PY['match']=FY_PY['py_job'].map(Final_Mapped_dict_PYjob['match'])
    FY_PY['Client_Match']=''
    FY_PY = FY_PY.rename(columns={'py_job':'job_number'})
    Final_Mapped['Match_column']=Final_Mapped['py_job'].map(str)+Final_Mapped['client_no'].map(str)
    FY_PY['Match_column']=FY_PY['job_number'].map(str)+FY_PY['client_no'].map(str)
    FY_PY_dict=FY_PY.set_index('Match_column').to_dict()
    Final_Mapped_dict=Final_Mapped.set_index('Match_column').to_dict()
    Final_Mapped['py_pieces']=Final_Mapped['Match_column'].map(FY_PY_dict['process_items_updated'])
    FY_PY['Client_Match']=FY_PY['Match_column'].map(Final_Mapped_dict['match'])
    FY_PY['Client_Match'][(FY_PY['match']=='Y') & (FY_PY['Client_Match']!='Y')] = 'Not in CY'
    FY_PY['match'][FY_PY['match']!='Y'] = 'N'
    FY_PY['Client_Match'][FY_PY['match']=='N'] = 'N'
    Interim=FY_PY[FY_PY['Client_Match']=='Not in CY']
    Interim=Interim.reset_index()
    del Interim['index']
    Interim = Interim.rename(columns={'job_number':'py_job','process_items':'py_pieces'})
    Interim['job_number']=Interim['py_job'].map(Final_Mapped_dict_PYjob['job_number'])
    Interim['process_items']=0
    Interim=Interim[['job_number','client_no','py_pieces','process_items','client_name','cusip6','py_job','match','Client_Match']]
    Final_Mapped=Final_Mapped.append(Interim)
    Final_Mapped=Final_Mapped.reset_index()
    del Final_Mapped['index']
    Final_Mapped['match'][Final_Mapped['match']!='Y'] = 'N'
    Final_Mapped['Cy_pieces_old'].fillna(0, inplace=True)
#Final_Mapped = Final_Mapped[[ 'job_number','jobtype', 'cusip6','client_no', 'client_name','issuer_name', 'exch_name', 'status', 'service', 'ml_date', 'record_date','jbrc_location','process_items', 'mailed_items', 'evs_consol_items', 'wrap_items', 'householding_items', 'nomail_items', 'email_items', 'items','F/C','match','py_ml_date','py_pieces','py_job','Cy_pieces_old','py_pieces_old']]
    Final_Mapped = Final_Mapped[[ 'job_number','jobtype', 'cusip6','client_no', 'client_name','issuer_name', 'exch_name', 'status', 'service', 'ml_date', 'record_date','jbrc_location','process_items','process_items_old_client', 'mailed_items', 'evs_consol_items', 'wrap_items', 'householding_items', 'nomail_items', 'email_items', 'items','F/C','match','py_pieces','py_job','Cy_pieces_old','py_pieces_old','Client_Match']]
    Final_Mapped_cusip6_CY = pd.pivot_table(Final_Mapped,index=["cusip6","job_number"],values=["process_items"],aggfunc=[np.sum],fill_value=0).reset_index()
    Final_Mapped_cusip6_CY.columns = Final_Mapped_cusip6_CY.columns.get_level_values(0)
    Final_Mapped_cusip6_CY = Final_Mapped_cusip6_CY.rename(columns={'sum':'new_process_items'})
    Final_Mapped_cusip6_CY_dict=Final_Mapped_cusip6_CY.set_index('cusip6').to_dict()
    Final_Mapped['new_process_items_cusip']=Final_Mapped['cusip6'].map(Final_Mapped_cusip6_CY_dict['new_process_items'])
    Final_Mapped_cusip6_PY= pd.pivot_table(Final_Mapped,index=["cusip6","job_number"],values=["py_pieces"],aggfunc=[np.sum],fill_value=0).reset_index()
    Final_Mapped_cusip6_PY.columns = Final_Mapped_cusip6_PY.columns.get_level_values(0)
    Final_Mapped_cusip6_PY = Final_Mapped_cusip6_PY.rename(columns={'sum':'new_py_pieces'})
#    Final_Mapped_cusip6_PY_dict=Final_Mapped_cusip6_PY.set_index('cusip6').to_dict()
    Final_Mapped['new_py_pieces_cusip']=Final_Mapped['cusip6'].map(Final_Mapped_cusip6_PY['new_py_pieces'])
    Final_Mapped_cusip6_CY['merge_column']=Final_Mapped_cusip6_CY['job_number']+Final_Mapped_cusip6_CY['cusip6'].map(str)
    Final_Mapped_cusip6_PY['merge_column']=Final_Mapped_cusip6_PY['job_number']+Final_Mapped_cusip6_PY['cusip6'].map(str)
    Cusip_Result['merge_column']=Cusip_Result['job_number']+Cusip_Result['cusip6'].map(str)
    Final_Mapped_cusip6_dict_CY=Final_Mapped_cusip6_CY.set_index('merge_column').to_dict()
    Final_Mapped_cusip6_dict_PY=Final_Mapped_cusip6_PY.set_index('merge_column').to_dict()
    Cusip_Result['new_process_items']=Cusip_Result['merge_column'].map(Final_Mapped_cusip6_dict_CY['new_process_items'])
    Cusip_Result['new_py_pieces']=Cusip_Result['merge_column'].map(Final_Mapped_cusip6_dict_PY['new_py_pieces'])
    Cusip_Result['old_process_items']=Cusip_Result['process_items']
    del Cusip_Result['process_items']
    Cusip_Result['process_items']=Cusip_Result['new_process_items']
    del Cusip_Result['new_process_items']
    Cusip_Result['old_py_pieces']=Cusip_Result['py_pieces']
    del Cusip_Result['py_pieces']
    Cusip_Result['py_pieces']=Cusip_Result['new_py_pieces']
    del Cusip_Result['new_py_pieces']
    Final_Mapped['old_process_items_cusip']=Final_Mapped['Cy_pieces_old']
    del Final_Mapped['Cy_pieces_old']
    Final_Mapped['Cy_pieces_old']=Final_Mapped['new_process_items_cusip']
    del Final_Mapped['new_process_items_cusip']
    Cusip_Result = Cusip_Result[[ u'job_number', u'jobtype', u'cusip6', u'issuer_name', u'exch_name', u'status', u'service', u'ml_date', u'record_date', u'jbrc_location', u'process_items',u'old_process_items', u'mailed_items', u'evs_consol_items', u'wrap_items', u'householding_items', u'nomail_items', u'email_items', u'items', u'F/C', u'match', u'py_ml_date', u'py_pieces',u'old_py_pieces', u'py_job']]
    Cusip_Map=deepcopy(Cusip_Result)
    Cusip_data=pd.read_csv(Interim_File_Path + r'\g_f_2019-05-31.txt',sep='|')
    Cusip_data['cusip6'] = Cusip_data['Primary Cusip'].astype(str).str[:6]
    Cusip_data = Cusip_data[['Industry','Sector','Last Annual Revenue','Last Annual Net Income','Last Annual EPS','Last Annual Total Assets','Number of Employees','Total Shares Outstanding','Primary Cusip','cusip6']]
    Cusip_Merge= pd.merge(Cusip_Map,Cusip_data,on='cusip6',how='left')
#    Temp_client=deepcopy(Final_Mapped)
    Cusip_Map_client=Final_Mapped
    Cusip_Merge_client= pd.merge(Cusip_Map_client,Cusip_data,on='cusip6',how='left')
    
    print ('client level analysis is successful for '+Logic)
    #==============================================================================
    #SUMMARY BUILDING for CUSIP level Report
    #==============================================================================
    print ('SUMMARY BUILDING for CUSIP level Report initiated for '+Logic)
    equity_data = deepcopy(Cusip_Result)
    #equity_data = pd.read_excel(r'C:\Data Science\SRG\EQ\CY-05-08\AfterMapping\EQ__AfterMapping_CY-05-08 16-58-03Interm_step3.xlsx')
    equity_data=equity_data[equity_data['match']=='Y']
    
    equity_data=equity_data.reset_index()
    del equity_data['index']
    
    equity_data['merged'] = equity_data['cusip6'].astype(str).map(str.strip) 
#    + equity_data['issuer_name'].astype(str).map(str.strip)
    equity_data_group= equity_data.groupby(['merged']).sum().reset_index()
    equity_data_group_dict=equity_data_group.set_index('merged').to_dict()
    equity_data['merged_process_items']=equity_data['merged'].map(equity_data_group_dict['process_items'])

#    Cusip_Merge_Sector_pivot = pd.pivot_table(equity_data,index=["Sector"],values=["process_items","py_pieces"],aggfunc=[np.sum],fill_value=0).reset_index()

    equity_data_200kplus = equity_data[equity_data['merged_process_items'].map(int)>200000]
    equity_data_200kplus['merged_process_items']
    equity_data_50k_200k = equity_data[(equity_data['merged_process_items'].map(int)<=200000) & (equity_data['merged_process_items'].map(int)>=50000)]
    equity_data_50k_200k['merged_process_items']
    equity_data_50Kless = equity_data[equity_data['merged_process_items'].map(int)<50000]
    equity_data_50Kless['merged_process_items']
    #-------------------------------------200kplus-------------------------------------
        
    counter = len(equity_data_200kplus['merged'].unique())
    
    Cusip_Summary = pd.DataFrame(columns=['cusip','job_number','Issuer','FY PY','FY CY','SRG %'])
    ytd1 = pd.DataFrame(columns=['cusip','job_number','Issuer','FY PY','FY CY','SRG %'],index=np.arange(counter))
     
    equity_data_200kplus = equity_data_200kplus.reset_index()
    del equity_data_200kplus['index']
    
    counter = 0
    totalLengthMerged = len(equity_data_200kplus['merged'].unique())
    for merge in equity_data_200kplus['merged'].unique():
        print (counter,totalLengthMerged)
        ytd1['FY CY'][counter] = equity_data_200kplus[equity_data_200kplus['merged']==merge]['process_items'].sum()
        ytd1['Issuer'][counter] = equity_data_200kplus[equity_data_200kplus['merged']==merge].reset_index()['issuer_name'][0]
        ytd1['job_number'][counter] = equity_data_200kplus[equity_data_200kplus['merged']==merge].reset_index()['job_number'][0]
        ytd1['cusip'][counter] = equity_data_200kplus[equity_data_200kplus['merged']==merge].reset_index()['cusip6'][0]
        ytd1['FY PY'][counter] = equity_data_200kplus[equity_data_200kplus['merged']==merge]['py_pieces'].sum()
        
        counter += 1
    
    ytd1 = ytd1[ytd1['FY PY']!=0]     
    
    ytd1['SRG %'] = ((((ytd1['FY CY'].map(float)/ytd1['FY PY'].map(float))-1)*100).round(2)).map(str) + ' %'
    ytd1['% Contribution']=((((ytd1['FY CY'].map(float)-ytd1['FY PY'].map(float))/equity_data['py_pieces'].map(float).sum())*100).round(2)).map(str) + ' %'
    ytd1_total = pd.DataFrame(columns=['job_number','cusip','Issuer','FY PY','FY CY','SRG %'],index=np.arange(1))
    ytd1_total['FY PY'][0] = ytd1['FY PY'].sum()
    ytd1_total['FY CY'][0] = ytd1['FY CY'].sum()
    ytd1_total['SRG %']= str(round(((ytd1_total['FY CY'].map(float)[0]-ytd1_total['FY PY'].map(float)[0])/ytd1_total['FY PY'].map(float)[0])*100,2)) + ' %'
    ytd1_total['cusip'][0] = 'TOTAL'
    
    py200 = ytd1_total['FY PY'].map(float)[0]
    cy200 = ytd1_total['FY CY'].map(float)[0]
    ytd1_total['% Contribution'] = str(round(((ytd1_total['FY CY'].map(float)[0]-ytd1_total['FY PY'].map(float)[0])/equity_data['py_pieces'].map(float).sum())*100,2)) + ' %'
    percent200 = str(round(((ytd1_total['FY CY'].map(float)[0]-ytd1_total['FY PY'].map(float)[0])/ytd1_total['FY PY'].map(float)[0])*100,2)) + ' %'
    ytd1=ytd1.sort_values(by=['% Contribution'],ascending=False)
    blank_row = pd.DataFrame(columns=ytd1.columns,index=np.arange(1))
    
    Cusip_Summary = Cusip_Summary.append(ytd1)
    Cusip_Summary = Cusip_Summary.append(ytd1_total)
    Cusip_Summary = Cusip_Summary.append(blank_row)

    #------------------------------------------------------------------------------
    #
    #-----------------------------50k_200k----------------------------------------
    counter1 = len(equity_data_50k_200k['merged'].unique())
    ytd1 = pd.DataFrame(columns=['cusip','job_number','Issuer','FY PY','FY CY','SRG %'],index=np.arange(counter1))
    equity_data_50k_200k = equity_data_50k_200k.reset_index()
    del equity_data_50k_200k['index']

    counter = 0
    for num,merge in enumerate(equity_data_50k_200k['merged'].unique()):
        print(num,counter1)
        ytd1['FY CY'][counter] = equity_data_50k_200k[equity_data_50k_200k['merged']==merge]['process_items'].sum()
        ytd1['Issuer'][counter] = equity_data_50k_200k[equity_data_50k_200k['merged']==merge].reset_index()['issuer_name'][0]
        ytd1['job_number'][counter] = equity_data_50k_200k[equity_data_50k_200k['merged']==merge].reset_index()['job_number'][0]
        ytd1['cusip'][counter] = equity_data_50k_200k[equity_data_50k_200k['merged']==merge].reset_index()['cusip6'][0]
        ytd1['FY PY'][counter] = equity_data_50k_200k[equity_data_50k_200k['merged']==merge]['py_pieces'].sum()
        counter += 1
    
    ytd1 = ytd1[ytd1['FY PY']!=0]     
    ytd1['SRG %'] = ((((ytd1['FY CY'].map(float)/ytd1['FY PY'].map(float))-1)*100).round(2)).map(str) + ' %'
    ytd1['% Contribution']=((((ytd1['FY CY'].map(float)-ytd1['FY PY'].map(float))/equity_data['py_pieces'].map(float).sum())*100).round(2)).map(str) + ' %'
    ytd1_total = pd.DataFrame(columns=['cusip','job_number','Issuer','FY PY','FY CY','SRG %'],index=np.arange(1))
    ytd1_total['FY PY'][0] = ytd1['FY PY'].sum()
    ytd1_total['FY CY'][0] = ytd1['FY CY'].sum()
    ytd1_total['SRG %']= str(round(((ytd1_total['FY CY'].map(float)[0]-ytd1_total['FY PY'].map(float)[0])/ytd1_total['FY PY'].map(float)[0])*100,2)) + ' %'
    ytd1_total['cusip'][0] = 'TOTAL'
    ytd1_total['% Contribution'] = str(round(((ytd1_total['FY CY'].map(float)[0]-ytd1_total['FY PY'].map(float)[0])/equity_data['py_pieces'].map(float).sum())*100,2)) + ' %'
    py50_200 = ytd1_total['FY PY'].map(float)[0]
    cy50_200 = ytd1_total['FY CY'].map(float)[0]
    percent50_200 = str(round(((ytd1_total['FY CY'].map(float)[0]-ytd1_total['FY PY'].map(float)[0])/ytd1_total['FY PY'].map(float)[0])*100,2)) + ' %'
    ytd1=ytd1.sort_values(by=['% Contribution'],ascending=False)
    blank_row = pd.DataFrame(columns=ytd1.columns,index=np.arange(1))
    Cusip_Summary = Cusip_Summary.append(ytd1)
    Cusip_Summary = Cusip_Summary.append(ytd1_total)
    Cusip_Summary = Cusip_Summary.append(blank_row)
    
    #------------------------------------------------------------------------------
    #
    #---------------------------------50Kless----------------------------------
    
    counter1 = len(equity_data_50Kless['merged'].unique())
    ytd1 = pd.DataFrame(columns=['cusip','job_number','Issuer','FY PY','FY CY','SRG %'],index=np.arange(counter1))
     
     
    equity_data_50Kless = equity_data_50Kless.reset_index()
    del equity_data_50Kless['index']
    
    counter = 0
    for num,merge in enumerate(equity_data_50Kless['merged'].unique()):
        print(num,counter1)
        ytd1['FY CY'][counter] = equity_data_50Kless[equity_data_50Kless['merged']==merge]['process_items'].sum()
        ytd1['Issuer'][counter] = equity_data_50Kless[equity_data_50Kless['merged']==merge].reset_index()['issuer_name'][0]
        ytd1['job_number'][counter] = equity_data_50Kless[equity_data_50Kless['merged']==merge].reset_index()['job_number'][0]
        ytd1['cusip'][counter] = equity_data_50Kless[equity_data_50Kless['merged']==merge].reset_index()['cusip6'][0]
        ytd1['FY PY'][counter] = equity_data_50Kless[equity_data_50Kless['merged']==merge]['py_pieces'].sum()
        counter += 1
        
    ytd1['SRG %'] = ((((ytd1['FY CY'].map(float)/ytd1['FY PY'].map(float))-1)*100).round(2)).map(str) + ' %'
    ytd1['% Contribution']=((((ytd1['FY CY'].map(float)-ytd1['FY PY'].map(float))/equity_data['py_pieces'].map(float).sum())*100).round(2)).map(str) + ' %'
    ytd1_total = pd.DataFrame(columns=['job_number','cusip','Issuer','FY PY','FY CY','SRG %'],index=np.arange(1))
    ytd1_total['FY PY'][0] = ytd1['FY PY'].sum()
    ytd1_total['FY CY'][0] = ytd1['FY CY'].sum()
    ytd1_total['SRG %']= str(round(((ytd1_total['FY CY'].map(float)[0]-ytd1_total['FY PY'].map(float)[0])/ytd1_total['FY PY'].map(float)[0])*100,2)) + ' %'
    ytd1_total['% Contribution'] = str(round(((ytd1_total['FY CY'].map(float)[0]-ytd1_total['FY PY'].map(float)[0])/equity_data['py_pieces'].map(float).sum())*100,2)) + ' %'
    ytd1_total['cusip'][0] = 'TOTAL'
    py50 = ytd1_total['FY PY'].map(float)[0]
    cy50 = ytd1_total['FY CY'].map(float)[0]
    percent50 = str(round(((ytd1_total['FY CY'].map(float)[0]-ytd1_total['FY PY'].map(float)[0])/ytd1_total['FY PY'].map(float)[0])*100,2)) + ' %'
    ytd1=ytd1.sort_values(by=['% Contribution'],ascending=False)
    blank_row = pd.DataFrame(columns=ytd1.columns,index=np.arange(1))
    Cusip_Summary = Cusip_Summary.append(ytd1)
    Cusip_Summary = Cusip_Summary.append(ytd1_total)
    Cusip_Summary = Cusip_Summary.append(blank_row)
    Cusip_Summary = Cusip_Summary.reset_index(drop=True)
    Cusip_Summary['FY CY'][Cusip_Summary['FY CY'].astype(str).map(str.strip)!='nan'] = Cusip_Summary[Cusip_Summary['FY CY'].astype(str).map(str.strip)!='nan']['FY CY'].astype(int)
    Cusip_Summary['FY PY'][Cusip_Summary['FY PY'].astype(str).map(str.strip)!='nan'] = Cusip_Summary[Cusip_Summary['FY PY'].astype(str).map(str.strip)!='nan']['FY PY'].astype(int)
    Cusip_Summary['FY CY'][Cusip_Summary['FY CY'].astype(str).map(str.strip)!='nan'] = Cusip_Summary[Cusip_Summary['FY CY'].astype(str).map(str.strip)!='nan']['FY CY'].astype(int).apply(lambda x: "{:,}".format(x)) 
    Cusip_Summary['FY PY'][Cusip_Summary['FY PY'].astype(str).map(str.strip)!='nan'] = Cusip_Summary[Cusip_Summary['FY PY'].astype(str).map(str.strip)!='nan']['FY PY'].astype(int).apply(lambda x: "{:,}".format(x)) 
    #------------------------------------------------------------------------------
    topTable_Cusip = pd.DataFrame(columns=['Beneficial Proxy Matched (Based on Mail Date)','','1','2','3','4','5'],index=np.arange(9))
    topTable_Cusip['Beneficial Proxy Matched (Based on Mail Date)'][0] = Logic +' -'+ Month + ' to '+ Month +' Pieces'
    topTable_Cusip['Beneficial Proxy Matched (Based on Mail Date)'][1] = 'Mail Date ' + Month+' '+str(PY)
    topTable_Cusip['Beneficial Proxy Matched (Based on Mail Date)'][2] = 'ISSUER View'
    topTable_Cusip['Beneficial Proxy Matched (Based on Mail Date)'][3] = ''
    topTable_Cusip['Beneficial Proxy Matched (Based on Mail Date)'][4] = 'Summary'
    topTable_Cusip['Beneficial Proxy Matched (Based on Mail Date)'][5] = 'Total Processed Pcs. Greater Than 200K'
    topTable_Cusip['Beneficial Proxy Matched (Based on Mail Date)'][6] = 'Total Processed Pcs. 50K - 200K'
    topTable_Cusip['Beneficial Proxy Matched (Based on Mail Date)'][7] = 'Total Processed Pcs. Less Than 50K'
    topTable_Cusip['Beneficial Proxy Matched (Based on Mail Date)'][8] = 'Total'
    topTable_Cusip['2'][3] = 'Processed Pieces'
    topTable_Cusip['3'][4] = PY
    topTable_Cusip['3'][5] = "{:,}".format(int(py200))
    topTable_Cusip['3'][6] = "{:,}".format(int(py50_200))
    topTable_Cusip['3'][7] = "{:,}".format(int(py50))
    topTable_Cusip['3'][8] = "{:,}".format(int(py200 + py50_200 + py50))
    topTable_Cusip['2'][4] = CY
    topTable_Cusip['2'][5] = "{:,}".format(int(cy200))
    topTable_Cusip['2'][6] = "{:,}".format(int(cy50_200))
    topTable_Cusip['2'][7] = "{:,}".format(int(cy50))
    topTable_Cusip['2'][8] = "{:,}".format(int(cy200 + cy50_200 + cy50))
    topTable_Cusip['4'][4] = 'SRG %'
    topTable_Cusip['4'][5] = percent200
    topTable_Cusip['4'][6] = percent50_200
    topTable_Cusip['4'][7] = percent50
    topTable_Cusip['4'][8] = str(round((((float(cy200 + cy50_200 + cy50)/float(py200 + py50_200 + py50))-1)*100),2)) + ' %'
    topTable_Cusip['5'][4] = '% Contribution'
    topTable_Cusip['5'][5] = str(round(((float(cy200)-float(py200))/float(py200 + py50_200 + py50))*100,2))+ ' %'
    topTable_Cusip['5'][6] = str(round(((float(cy50_200)-float(py50_200))/float(py200 + py50_200 + py50))*100,2))+ ' %'
    topTable_Cusip['5'][7] = str(round(((float(cy50)-float(py50))/float(py200 + py50_200 + py50))*100,2))+ ' %'
    topTable_Cusip['5'][8] = str(round((((float(cy200 + cy50_200 + cy50)/float(py200 + py50_200 + py50))-1)*100),2)) + ' %'
    Cusip_total = pd.DataFrame(columns=Cusip_Summary.columns,index=np.arange(1))
    Cusip_total['cusip'][0] = 'Total Processed Pieces'
    Cusip_total['FY PY'][0] = topTable_Cusip['3'][8]
    Cusip_total['FY CY'][0] = topTable_Cusip['2'][8]
    Cusip_total['SRG %'][0] = topTable_Cusip['4'][8]
    Cusip_total['% Contribution'][0] = topTable_Cusip['5'][8]
    Cusip_Summary = Cusip_Summary.append(Cusip_total)
    Cusip_Summary=Cusip_Summary.rename(columns={'cusip':'Cusip','FY PY':'FY '+str(PY),'FY CY':'FY '+str(CY),'job_number':'Job Number'})
    Cusip_Summary=Cusip_Summary[['Cusip','Job Number','Issuer','FY '+str(CY),'FY '+str(PY),'SRG %','% Contribution']]
    topTable_Cusip = topTable_Cusip.rename(columns={'1':'','2':'','3':'','4':'','5':''})

    print ('SUMMARY BUILDING for CUSIP level Report successful for '+Logic)

    #==============================================================================
    # SUMMARY BUILDING for client level Report
    #==============================================================================
    print ('SUMMARY BUILDING for client level Report initiated for '+Logic)

    equity_data = deepcopy(Final_Mapped)
        #equity_data = pd.read_excel(r'C:\Data Science\SRG\EQ\2018-05-08\AfterMapping\EQ__AfterMapping_2018-05-08 16-58-03Interm_step3.xlsx')
    equity_data=equity_data[equity_data['match']=='Y']
    
    equity_data['merged_cusip'] = equity_data['cusip6'].astype(str).map(str.strip) 
    equity_data_group= equity_data.groupby(['merged_cusip']).sum().reset_index()
    equity_data_group_dict=equity_data_group.set_index('merged_cusip').to_dict()
    equity_data['merged_process_items']=equity_data['merged_cusip'].map(equity_data_group_dict['process_items'])

    equity_data['py_pieces']=equity_data['py_pieces'].convert_objects(convert_numeric=True).fillna(0)
    equity_data['merged'] = equity_data['client_no'].astype(str).map(str.strip) + equity_data['client_name'].astype(str).map(str.strip)
    equity_data_200kplus = equity_data[equity_data['merged_process_items'].map(int)>200000]
    
    equity_data_50k_200k = equity_data[(equity_data['merged_process_items'].map(int)<=200000) & (equity_data['merged_process_items'].map(int)>=50000)]
    
    equity_data_50Kless = equity_data[equity_data['merged_process_items'].map(int)<50000]
    
    #-------------------------------------200kplus-------------------------------------
    
        
    counter1 = len(equity_data_200kplus['merged'].unique())
        
    Summary_client = pd.DataFrame(columns=['client_no','client_name','FY PY','FY CY','SRG %'])
    ytd1 = pd.DataFrame(columns=['client_no','client_name','FY PY','FY CY','SRG %'],index=np.arange(counter1))
    equity_data_200kplus = equity_data_200kplus.reset_index()
    del equity_data_200kplus['index']
    
    counter = 0
    for num,merge in enumerate(equity_data_200kplus['merged'].unique()):
        print(num,counter1)
        ytd1['FY CY'][counter] = equity_data_200kplus[equity_data_200kplus['merged']==merge]['process_items'].sum()
        ytd1['client_name'][counter] = equity_data_200kplus[equity_data_200kplus['merged']==merge].reset_index()['client_name'][0]
        ytd1['client_no'][counter] = equity_data_200kplus[equity_data_200kplus['merged']==merge].reset_index()['client_no'][0]
        ytd1['FY PY'][counter] = equity_data_200kplus[equity_data_200kplus['merged']==merge]['py_pieces'].sum()
        counter += 1
        
    ytd1['SRG %'] = ((((ytd1['FY CY'].map(float)/ytd1['FY PY'].map(float))-1)*100).round(2)).map(str) + ' %'
    ytd1['% Contribution']=((((ytd1['FY CY'].map(float)-ytd1['FY PY'].map(float))/equity_data['py_pieces'].map(float).sum())*100).round(2)).map(str) + ' %'
    ytd1_total = pd.DataFrame(columns=['client_no','client_name','FY PY','FY CY','SRG %'],index=np.arange(1))
    ytd1_total['FY PY'][0] = ytd1['FY PY'].sum()
    ytd1_total['FY CY'][0] = ytd1['FY CY'].sum()
    ytd1_total['SRG %']= str(round(((ytd1_total['FY CY'].map(float)[0]-ytd1_total['FY PY'].map(float)[0])/ytd1_total['FY PY'].map(float)[0])*100,2)) + ' %'
    ytd1_total['client_no'][0] = 'TOTAL'
    py200 = ytd1_total['FY PY'].map(float)[0]
    cy200 = ytd1_total['FY CY'].map(float)[0]
    ytd1_total['% Contribution'] = str(round(((ytd1_total['FY CY'].map(float)[0]-ytd1_total['FY PY'].map(float)[0])/equity_data['py_pieces'].map(float).sum())*100,2)) + ' %'
    percent200 = str(round(((ytd1_total['FY CY'].map(float)[0]-ytd1_total['FY PY'].map(float)[0])/ytd1_total['FY PY'].map(float)[0])*100,2)) + ' %'
    ytd1=ytd1.sort_values(by=['% Contribution'],ascending=False)
    blank_row = pd.DataFrame(columns=ytd1.columns,index=np.arange(1))
    Summary_client = Summary_client.append(ytd1)
    Summary_client = Summary_client.append(ytd1_total)
    Summary_client = Summary_client.append(blank_row)
    
    #-----------------------------50k_200k----------------------------------------
        
    counter1 = len(equity_data_50k_200k['merged'].unique())
    
    ytd1 = pd.DataFrame(columns=['client_no','client_name','FY PY','FY CY','SRG %'],index=np.arange(counter1))
    equity_data_50k_200k = equity_data_50k_200k.reset_index()
    del equity_data_50k_200k['index']
    
    counter = 0
    for num,merge in enumerate(equity_data_50k_200k['merged'].unique()):
        print(num,counter1)
        ytd1['FY CY'][counter] = equity_data_50k_200k[equity_data_50k_200k['merged']==merge]['process_items'].sum()
        ytd1['client_name'][counter] = equity_data_50k_200k[equity_data_50k_200k['merged']==merge].reset_index()['client_name'][0]
        ytd1['client_no'][counter] = equity_data_50k_200k[equity_data_50k_200k['merged']==merge].reset_index()['client_no'][0]
        ytd1['FY PY'][counter] = equity_data_50k_200k[equity_data_50k_200k['merged']==merge]['py_pieces'].sum()
        counter += 1
    
    ytd1['SRG %'] = ((((ytd1['FY CY'].map(float)/ytd1['FY PY'].map(float))-1)*100).round(2)).map(str) + ' %'
    ytd1['% Contribution']=((((ytd1['FY CY'].map(float)-ytd1['FY PY'].map(float))/equity_data['py_pieces'].map(float).sum())*100).round(2)).map(str) + ' %'
    ytd1_total = pd.DataFrame(columns=['client_no','client_name','FY PY','FY CY','SRG %'],index=np.arange(1))
    ytd1_total['FY PY'][0] = ytd1['FY PY'].sum()
    ytd1_total['FY CY'][0] = ytd1['FY CY'].sum()
    ytd1_total['SRG %']= str(round(((ytd1_total['FY CY'].map(float)[0]-ytd1_total['FY PY'].map(float)[0])/ytd1_total['FY PY'].map(float)[0])*100,2)) + ' %'
    ytd1_total['client_no'][0] = 'TOTAL'
    py50_200 = ytd1_total['FY PY'].map(float)[0]
    cy50_200 = ytd1_total['FY CY'].map(float)[0]
    ytd1_total['% Contribution'] = str(round(((ytd1_total['FY CY'].map(float)[0]-ytd1_total['FY PY'].map(float)[0])/equity_data['py_pieces'].map(float).sum())*100,2)) + ' %'
    percent50_200 =str(round(((ytd1_total['FY CY'].map(float)[0]-ytd1_total['FY PY'].map(float)[0])/ytd1_total['FY PY'].map(float)[0])*100,2)) + ' %'
    ytd1=ytd1.sort_values(by=['% Contribution'],ascending=False)
    blank_row = pd.DataFrame(columns=ytd1.columns,index=np.arange(1))
    Summary_client = Summary_client.append(ytd1)
    Summary_client = Summary_client.append(ytd1_total)
    Summary_client = Summary_client.append(blank_row)    
    #---------------------------------50Kless----------------------------------
    
    counter1 = len(equity_data_50Kless['merged'].unique())
    
    ytd1 = pd.DataFrame(columns=['client_no','client_name','FY PY','FY CY','SRG %'],index=np.arange(counter))
    equity_data_50Kless = equity_data_50Kless.reset_index()
    del equity_data_50Kless['index']
    
    counter = 0
    for num,merge in enumerate(equity_data_50Kless['merged'].unique()):
        print(num,counter1)
        ytd1['FY CY'][counter] = equity_data_50Kless[equity_data_50Kless['merged']==merge]['process_items'].sum()
        ytd1['client_name'][counter] = equity_data_50Kless[equity_data_50Kless['merged']==merge].reset_index()['client_name'][0]
        ytd1['client_no'][counter] = equity_data_50Kless[equity_data_50Kless['merged']==merge].reset_index()['client_no'][0]
        ytd1['FY PY'][counter] = equity_data_50Kless[equity_data_50Kless['merged']==merge]['py_pieces'][equity_data_50Kless[equity_data_50Kless['merged']==merge]['py_pieces']!=0].sum()
        counter += 1
    
    ytd1['SRG %'] = ((((ytd1['FY CY'].map(float)/ytd1['FY PY'].map(float))-1)*100).round(2)).map(str) + ' %'
    ytd1['% Contribution']=((((ytd1['FY CY'].map(float)-ytd1['FY PY'].map(float))/equity_data['py_pieces'].map(float).sum())*100).round(2)).map(str) + ' %'
    ytd1_total = pd.DataFrame(columns=['client_no','client_name','FY PY','FY CY','SRG %'],index=np.arange(1))
    ytd1_total['FY PY'][0] = ytd1['FY PY'].sum()
    ytd1_total['FY CY'][0] = ytd1['FY CY'].sum()
    ytd1_total['SRG %']= str(round(((ytd1_total['FY CY'].map(float)[0]-ytd1_total['FY PY'].map(float)[0])/ytd1_total['FY PY'].map(float)[0])*100,2)) + ' %'
    ytd1_total['client_no'][0] = 'TOTAL'
    py50 = ytd1_total['FY PY'].map(float)[0]
    cy50 = ytd1_total['FY CY'].map(float)[0]
    ytd1_total['% Contribution'] = str(round(((ytd1_total['FY CY'].map(float)[0]-ytd1_total['FY PY'].map(float)[0])/equity_data['py_pieces'].map(float).sum())*100,2)) + ' %'
    percent50 =  str(round(((ytd1_total['FY CY'].map(float)[0]-ytd1_total['FY PY'].map(float)[0])/ytd1_total['FY PY'].map(float)[0])*100,2)) + ' %'
    ytd1=ytd1.sort_values(by=['% Contribution'],ascending=False)
    blank_row = pd.DataFrame(columns=ytd1.columns,index=np.arange(1))
    Summary_client = Summary_client.append(ytd1)
    Summary_client = Summary_client.append(ytd1_total)
    Summary_client = Summary_client.append(blank_row)
    topTable_client = pd.DataFrame(columns=['Beneficial Proxy Matched (Based on Mail Date)','','1','2','3','4'],index=np.arange(9))
    topTable_client['Beneficial Proxy Matched (Based on Mail Date)'][0] = 'MTD - '+ Month + ' to '+ Month +' Pieces'
    topTable_client['Beneficial Proxy Matched (Based on Mail Date)'][1] = 'Mail Date ' + Month+' '+str( PY)
    topTable_client['Beneficial Proxy Matched (Based on Mail Date)'][2] = 'Client View'
    topTable_client['Beneficial Proxy Matched (Based on Mail Date)'][3] = ''
    topTable_client['Beneficial Proxy Matched (Based on Mail Date)'][4] = 'Summary'
    topTable_client['Beneficial Proxy Matched (Based on Mail Date)'][5] = 'Total Processed Pcs. Greater Than 200K'
    topTable_client['Beneficial Proxy Matched (Based on Mail Date)'][6] = 'Total Processed Pcs. 50K - 200K'
    topTable_client['Beneficial Proxy Matched (Based on Mail Date)'][7] = 'Total Processed Pcs. Less Than 50K'
    topTable_client['Beneficial Proxy Matched (Based on Mail Date)'][8] = 'Total'
    topTable_client['1'][3] = 'Processed Pieces'
    topTable_client['1'][4] = CY
    topTable_client['1'][5] = "{:,}".format(int(cy200))
    topTable_client['1'][6] = "{:,}".format(int(cy50_200))
    topTable_client['1'][7] = "{:,}".format(int(cy50))
    topTable_client['1'][8] = "{:,}".format(int(cy200 + cy50_200 + cy50))
    topTable_client['2'][4] = PY
    topTable_client['2'][5] = "{:,}".format(int(py200))
    topTable_client['2'][6] = "{:,}".format(int(py50_200))
    topTable_client['2'][7] = "{:,}".format(int(py50))
    topTable_client['2'][8] = "{:,}".format(int(py200 + py50_200 + py50))
    topTable_client['3'][4] = 'SRG %'
    topTable_client['3'][5] = percent200
    topTable_client['3'][6] = percent50_200
    topTable_client['3'][7] = percent50
    topTable_client['3'][8] = str(round((((float(cy200 + cy50_200 + cy50)/float(py200 + py50_200 + py50))-1)*100),2)) + ' %'
    topTable_client['4'][4] = '% Contribution'
    topTable_client['4'][5] = str(round(((float(cy200)-float(py200))/float(py200 + py50_200 + py50))*100,2))+ ' %'
    topTable_client['4'][6] = str(round(((float(cy50_200)-float(py50_200))/float(py200 + py50_200 + py50))*100,2))+ ' %'
    topTable_client['4'][7] = str(round(((float(cy50)-float(py50))/float(py200 + py50_200 + py50))*100,2))+ ' %'
    topTable_client['4'][8] = str(round((((float(cy200 + cy50_200 + cy50)/float(py200 + py50_200 + py50))-1)*100),2)) + ' %'
    Total_Client = pd.DataFrame(columns=Summary_client.columns,index=np.arange(1))
    Total_Client['client_no'][0] = 'Total Processed Pieces'
    Total_Client['FY PY'][0] = topTable_client['2'][8]
    Total_Client['FY CY'][0] = topTable_client['1'][8]
    Total_Client['SRG %'][0] = topTable_client['3'][8]
    Total_Client['% Contribution'][0] = topTable_client['4'][8]
    Summary_client = Summary_client.append(Total_Client)
    Summary_client=Summary_client.rename(columns={'client_no':'Client No','FY PY':'FY '+str(PY),'FY CY':'FY '+str(CY),'client_name':'Client Name'})
    Summary_client=Summary_client[['Client No','Client Name','FY '+str(CY),'FY '+str(PY),'SRG %','% Contribution']]
    topTable_client = topTable_client.rename(columns={'1':'','2':'','3':'','4':''})
    
    print ('SUMMARY BUILDING for client level Report successful')
    # =============================================================================
    # Sector Report Summary
    # =============================================================================
    print ('SUMMARY BUILDING for Sector level Report initiated')
    
    topTable_Sector = pd.DataFrame(columns=['Beneficial Proxy Matched (Based on Mail Date)','','1','2','3'],index=np.arange(9))
    topTable_Sector['Beneficial Proxy Matched (Based on Mail Date)'][0] = Logic +' -'+ Month + ' to '+ Month +' Pieces'
    topTable_Sector['Beneficial Proxy Matched (Based on Mail Date)'][1] = 'Mail Date ' + Month+' '+str( PY)
    topTable_Sector['Beneficial Proxy Matched (Based on Mail Date)'][2] = 'Sector View'
    topTable_Sector['Beneficial Proxy Matched (Based on Mail Date)'][3] = ''
    topTable_Sector = topTable_Sector.rename(columns={'1':'','2':'','3':''})
    Cusip_Merge_Matched=deepcopy(Cusip_Merge)
    Cusip_Merge_Matched=Cusip_Merge_Matched[Cusip_Merge_Matched['match']=='Y']
    Cusip_Merge_Matched["Sector"].fillna("Unidentified", inplace = True) 
    Cusip_Merge_Sector_pivot = pd.pivot_table(Cusip_Merge_Matched,index=["Sector"],values=["process_items","py_pieces"],aggfunc=[np.sum],fill_value=0).reset_index()
    Cusip_Merge_Sector_pivot.columns = Cusip_Merge_Sector_pivot.columns.get_level_values(0)
    Cusip_Merge_Sector_pivot = Cusip_Merge_Sector_pivot.rename(columns={'sum':'new_py_pieces'})
    Cusip_Merge_Sector_pivot.columns=['Sector','CY Pieces','PY Pieces']
    Cusip_Merge_Sector_pivot['SRG %']=((((Cusip_Merge_Sector_pivot['CY Pieces'].map(float)/Cusip_Merge_Sector_pivot['PY Pieces'].map(float))-1)*100).round(2)).map(str) + ' %'
    Cusip_Merge_Sector_pivot['% Contribution'] = ((((Cusip_Merge_Sector_pivot['CY Pieces'].map(float)-Cusip_Merge_Sector_pivot['PY Pieces'].map(float))/Cusip_Merge_Sector_pivot['PY Pieces'].sum())*100).round(2)).map(str) + ' %'
    Cusip_Merge_Sector_pivot_total = pd.DataFrame(columns=['Sector', 'CY Pieces', 'PY Pieces', 'SRG %', '% Contribution'],index=np.arange(1))
    Cusip_Merge_Sector_pivot_total['Sector'][0] = 'TOTAL'
    Cusip_Merge_Sector_pivot_total['CY Pieces'][0] = Cusip_Merge_Sector_pivot['CY Pieces'].sum()
    Cusip_Merge_Sector_pivot_total['PY Pieces'][0] = Cusip_Merge_Sector_pivot['PY Pieces'].sum()
    Cusip_Merge_Sector_pivot_total['SRG %'] = str((((Cusip_Merge_Sector_pivot['CY Pieces'].sum()/Cusip_Merge_Sector_pivot['PY Pieces'].sum()-1)*100).round(2))) + ' %'
    Cusip_Merge_Sector_pivot_total['% Contribution'][0] = str((((Cusip_Merge_Sector_pivot['CY Pieces'].sum()/Cusip_Merge_Sector_pivot['PY Pieces'].sum()-1)*100).round(2))) + ' %'
    Cusip_Merge_Sector_pivot=Cusip_Merge_Sector_pivot.append(Cusip_Merge_Sector_pivot_total)
        
    print ('SUMMARY BUILDING for Sector level Report successful for '+Logic)

    # =============================================================================
    # Write Data into File.
    # =============================================================================
    print ('Writing the data into the file for ' +Logic)

    Report = TodaysDirecotry_Cusip+'\Report\\'
    MakeDirectory(Report)
    File_Name = Report+'EQ_Cusip_Report_'+ Logic +'_'+str(Month_To_Be_Run) +'_'+str(CY) +'.xlsx'
    writer = pd.ExcelWriter(File_Name, engine = 'xlsxwriter')
    Cusip_Result.to_excel(writer, sheet_name = 'Cusip Data',index=False)
    print('Cusip Data written')
    Cusip_Merge.to_excel(writer, sheet_name = 'Cusip Details Data',index=False)
    print('Cusip Details Data written')
    topTable_Cusip.to_excel(writer, sheet_name = 'Report',index=False)
    print('Report 1 written')
    Cusip_Summary.to_excel(writer, sheet_name='Report',index=False,startrow=11)
    print('Report 2 written')
    Final_Mapped.to_excel(writer, sheet_name = 'Client Data',index=False)
    print('Client Data written')
    Cusip_Merge_client.to_excel(writer, sheet_name = 'Client Details Data',index=False)
    print('Client Details Data written')
    topTable_client.to_excel(writer, sheet_name = 'Client_Report',index=False)
    print('Client_Report 1 written')
    Summary_client.to_excel(writer, sheet_name='Client_Report',index=False,startrow=11)
    print('Client_Report 2 written')
#   writer_sector = pd.ExcelWriter(fn, engine = 'xlsxwriter')
    topTable_Sector.to_excel(writer, sheet_name = 'Sector_Report',index=False)
    print('Sector_Report 1 written')
    Cusip_Merge_Sector_pivot.to_excel(writer, sheet_name='Sector_Report',index=False,startrow=6)
    print('Sector_Report 2 written')
    writer.save()
    writer.close()
    
    print("Data is successfully written into file for "+Logic)

    # =============================================================================
    # Formatting the Cusip View
    # =============================================================================
    print("Formatting the Cusip view for "+Logic)
    wb = openpyxl.load_workbook(File_Name)
    sheets=wb.get_sheet_names()
    Cusip_Report = wb.get_sheet_by_name(sheets[2])
    
    for Whole_Sheet in Cusip_Report.iter_rows(min_row=1, max_col=3, max_row=1000):
        for cell in Whole_Sheet:
            cell.font = Font(name='Arial', size=10)
    
    for A6_G10 in Cusip_Report.iter_rows(min_row=6, max_col=7, max_row=10):
        for cell1 in A6_G10:
            cell1.font = Font(name='Arial', size=10, color="000000",bold=True)
            cell1.alignment = Alignment(vertical='bottom')
            cell1.alignment = Alignment(horizontal='left')
            
    thin_top_border = Border(top=Side(style='thin'))
    for A10_G10 in Cusip_Report.iter_rows(min_row=10, max_col=7, max_row=10):     
        for cell2 in A10_G10:
            cell2.border= thin_top_border
            
    D5=Cusip_Report['D5']
    D5.font = Font(name='Arial', size=10, color="000000",bold=True)
    Cusip_Report.merge_cells('D5:E5')
    
    for A6_G6 in Cusip_Report.iter_rows(min_row=6, max_col=7, max_row=6):
        for cell6 in A6_G6:
            cell6.font = Font(name='Arial', size=10, color="FFFFFF",bold=True,underline='single')
            cell6.alignment = Alignment(horizontal='left',wrap_text = False)
            cell6.fill = PatternFill("solid", fgColor="1F497D")
    
    
    thin_bottom_border = Border(bottom=Side(style='thin'))
    for A12_G12 in Cusip_Report.iter_rows(min_row=12, max_col=7, max_row=12):
        for cell7 in A12_G12:
            cell7.font = Font(name='Arial', size=10, color="000000",bold=True)
            cell7.alignment = Alignment(horizontal='left',wrap_text = False)
            cell7.border= thin_bottom_border
      
    No_border = Border()
    for A1_G1 in Cusip_Report.iter_rows(min_row=1, max_col=7, max_row=1):
        for cell8 in A1_G1:
            cell8.font = Font(name='Arial', size=10, color="000000",bold=True)
            cell8.alignment = Alignment(horizontal='left',wrap_text = False)
            cell8.border= No_border
    
    for C13_D10000 in Cusip_Report.iter_rows(min_row=13,min_col=4, max_col=7, max_row=10000):
        for cell9 in C13_D10000:
            cell9.font = Font(name='Arial', size=10)
            cell9.alignment = Alignment(horizontal='right',wrap_text = False)
            
    for A1_A1000 in Cusip_Report.iter_rows(min_row=1, max_col=1, max_row=10000):
        for cell3 in A1_A1000:
            if (cell3.value=='TOTAL') | (cell3.value=='Total Processed Pieces') | (cell3.value=='Cusip') |(cell3.value=='Total') :
                for specific_rows in Cusip_Report.iter_rows(min_row=cell3.row, max_col=7,max_row=cell3.row):
                    for cell4 in specific_rows:
                        cell4.font = Font(name='Arial', size=10, color="FFFFFF",bold=True)
                        cell4.alignment = Alignment(horizontal='left',wrap_text = False)
                        cell4.fill = PatternFill("solid", fgColor="1F497D")
    
    for column_cells in Cusip_Report.iter_rows(min_row=13, min_col=3, max_col=4,max_row=1000):
        Cusip_Report.column_dimensions[column_cells[0].column].width = 50
                        
    ListofBlankRows=[]    
    Loopcounter=0
    for A1_A10000 in Cusip_Report.iter_rows(min_row=13,max_row=10000, max_col=1):
        for cell10 in A1_A10000:
            if Loopcounter <3:
                if cell10.value==None:
                    Loopcounter+=1
                    ListofBlankRows.append(cell10.row)
                    
    FirstRowGreaterThan200K=13
    LastRowGreaterThan200K=ListofBlankRows[0]-2
    FirstRow50K_200K=ListofBlankRows[0]+1
    LastRow50K_200K=ListofBlankRows[1]-2
    FirstRowLessThan50K=ListofBlankRows[1]+1
    LastRowLessThan50K=ListofBlankRows[2]-2
    
    if (LastRowGreaterThan200K-FirstRowGreaterThan200K)>20:
        for idx in range(FirstRowGreaterThan200K+20, LastRowGreaterThan200K-5):
            Cusip_Report.row_dimensions[idx].hidden = True
        for Last5Rowsgreater200K in Cusip_Report.iter_rows(min_row=LastRowGreaterThan200K-5, min_col=1,max_col=7, max_row=LastRowGreaterThan200K): 
            for idx4 in Last5Rowsgreater200K:
                idx4.font = Font(name='Arial', size=10,color="ff0000")         
    if (LastRow50K_200K-FirstRow50K_200K)>20:
        for idx2 in range(FirstRow50K_200K+20, LastRow50K_200K-5):
            Cusip_Report.row_dimensions[idx2].hidden = True
        for Last5Rows50K_200K in Cusip_Report.iter_rows(min_row=LastRow50K_200K-5, min_col=1,max_col=7, max_row=LastRow50K_200K): 
            for idx5 in Last5Rows50K_200K:
                idx5.font = Font(name='Arial', size=10,color="ff0000")    
    
    if (LastRowLessThan50K-FirstRowLessThan50K)>20:
        for idx3 in range(FirstRowLessThan50K+20, LastRowLessThan50K-5):
            Cusip_Report.row_dimensions[idx3].hidden = True
        for Last5RowsLess50K in Cusip_Report.iter_rows(min_row=LastRowLessThan50K-5, min_col=1,max_col=7, max_row=LastRowLessThan50K): 
            for idx6 in Last5RowsLess50K:
                idx6.font = Font(name='Arial', size=10,color="ff0000")
                
                
    for Whole_Sheet_number in Cusip_Report.iter_rows(min_row=7,min_col=4, max_col=5, max_row=10000):
        for cell12 in Whole_Sheet_number:
            cell12.number_format= '#,##0'
            
            
    for Whole_Sheet_number2 in Cusip_Report.iter_rows(min_row=7,min_col=6, max_col=7, max_row=10000):
        for cell13 in Whole_Sheet_number2:
            cell13.number_format= '0.00%'
            
    Cusip_Report.freeze_panes = 'A13'
    print("Formatting successful the Cusip view for "+Logic)

    # =============================================================================
    # Formatting the Client Report
    # =============================================================================
    print("Formatting the client view for "+Logic)

    Client_Report = wb.get_sheet_by_name(sheets[5])
    
    for Whole_Sheet in Client_Report.iter_rows(min_row=1, max_col=3, max_row=10000):
        for cell in Whole_Sheet:
            cell.font = Font(name='Arial', size=10)
    
    for A6_G10 in Client_Report.iter_rows(min_row=6, max_col=7, max_row=10):
        for cell1 in A6_G10:
            cell1.font = Font(name='Arial', size=10, color="000000",bold=True)
            cell1.alignment = Alignment(vertical='bottom')
            cell1.alignment = Alignment(horizontal='left')
            
    thin_top_border = Border(top=Side(style='thin'))
    for A10_G10 in Client_Report.iter_rows(min_row=10, max_col=7, max_row=10):     
        for cell2 in A10_G10:
            cell2.border= thin_top_border
            
    D5=Client_Report['D5']
    D5.font = Font(name='Arial', size=10, color="000000",bold=True)
    Client_Report.merge_cells('D5:E5')
    
    for A6_G6 in Client_Report.iter_rows(min_row=6, max_col=6, max_row=6):
        for cell6 in A6_G6:
            cell6.font = Font(name='Arial', size=10, color="FFFFFF",bold=True,underline='single')
            cell6.alignment = Alignment(horizontal='left',wrap_text = False)
            cell6.fill = PatternFill("solid", fgColor="1F497D")
    
    
    thin_bottom_border = Border(bottom=Side(style='thin'))
    for A12_G12 in Client_Report.iter_rows(min_row=12, max_col=7, max_row=12):
        for cell7 in A12_G12:
            cell7.font = Font(name='Arial', size=10, color="000000",bold=True)
            cell7.alignment = Alignment(horizontal='left',wrap_text = False)
            cell7.border= thin_bottom_border
    
    No_border = Border()
    for A1_G1 in Client_Report.iter_rows(min_row=1, max_col=7, max_row=1):
        for cell8 in A1_G1:
            cell8.font = Font(name='Arial', size=10, color="000000",bold=True)
            cell8.alignment = Alignment(horizontal='left',wrap_text = False)
            cell8.border= No_border
    
    for C13_D10000 in Client_Report.iter_rows(min_row=13,min_col=4, max_col=7, max_row=10000):
        for cell9 in C13_D10000:
            cell9.font = Font(name='Arial', size=10)
            cell9.alignment = Alignment(horizontal='right',wrap_text = False)
            
    for A1_A10000 in Client_Report.iter_rows(min_row=1, max_col=1, max_row=10000):
        for cell3 in A1_A10000:
            if (cell3.value=='TOTAL') | (cell3.value=='Total Processed Pieces') | (cell3.value=='Client No')|(cell3.value=='Total'):
                for specific_rows in Client_Report.iter_rows(min_row=cell3.row, max_col=6,max_row=cell3.row):
                    for cell4 in specific_rows:
                        cell4.font = Font(name='Arial', size=10, color="FFFFFF",bold=True)
                        cell4.alignment = Alignment(horizontal='left',wrap_text = False)
                        cell4.fill = PatternFill("solid", fgColor="1F497D")
    
    for column_cells in Client_Report.iter_rows(min_row=13, min_col=2, max_col=2,max_row=10000):
        Client_Report.column_dimensions[column_cells[0].column].width = 50
                        
                       
    ListofBlankRows=[]    
    Loopcounter=0
    for A1_A10000 in Client_Report.iter_rows(min_row=13,max_row=10000, max_col=1):
        for cell10 in A1_A10000:
            if Loopcounter <3:
                if cell10.value==None:
                    Loopcounter+=1
                    ListofBlankRows.append(cell10.row)
    
    FirstRowGreaterThan200K=13
    LastRowGreaterThan200K=ListofBlankRows[0]-2
    FirstRow50K_200K=ListofBlankRows[0]+1
    LastRow50K_200K=ListofBlankRows[1]-2
    FirstRowLessThan50K=ListofBlankRows[1]+1
    LastRowLessThan50K=ListofBlankRows[2]-2
    
    if (LastRowGreaterThan200K-FirstRowGreaterThan200K)>20:
        for idx in range(FirstRowGreaterThan200K+20, LastRowGreaterThan200K-5):
            Client_Report.row_dimensions[idx].hidden = True
        for Last5Rowsgreater200K in Client_Report.iter_rows(min_row=LastRowGreaterThan200K-5, min_col=1,max_col=7, max_row=LastRowGreaterThan200K): 
            for idx4 in Last5Rowsgreater200K:
                idx4.font = Font(name='Arial', size=10,color="ff0000")         
    if (LastRow50K_200K-FirstRow50K_200K)>20:
        for idx2 in range(FirstRow50K_200K+20, LastRow50K_200K-5):
            Client_Report.row_dimensions[idx2].hidden = True
        for Last5Rows50K_200K in Client_Report.iter_rows(min_row=LastRow50K_200K-5, min_col=1,max_col=7, max_row=LastRow50K_200K): 
            for idx5 in Last5Rows50K_200K:
                idx5.font = Font(name='Arial', size=10,color="ff0000")    
    
    if (LastRowLessThan50K-FirstRowLessThan50K)>20:
        for idx3 in range(FirstRowLessThan50K+20, LastRowLessThan50K-5):
            Client_Report.row_dimensions[idx3].hidden = True
        for Last5RowsLess50K in Client_Report.iter_rows(min_row=LastRowLessThan50K-5, min_col=1,max_col=7, max_row=LastRowLessThan50K): 
            for idx6 in Last5RowsLess50K:
                idx6.font = Font(name='Arial', size=10,color="ff0000")
            
    Client_Report.freeze_panes = 'A13'
    
    for Whole_Sheet_number in Client_Report.iter_rows(min_row=7,min_col=3, max_col=4, max_row=10000):
        for cell12 in Whole_Sheet_number:
            cell12.number_format= '#,##0'
            
            
    for Whole_Sheet_number2 in Client_Report.iter_rows(min_row=7,min_col=5, max_col=6, max_row=10000):
        for cell13 in Whole_Sheet_number2:
            cell13.number_format= '0.00%'
            
    print("Formatting successful the client view for "+Logic)

    # =============================================================================
    # formatting Sector Report
    # =============================================================================
    print("Formatting the Sector view for "+Logic)

    Sector_Report = wb.get_sheet_by_name(sheets[6])
    
    for Whole_Sheet in Sector_Report.iter_rows(min_row=1, max_col=3, max_row=100):
        for cell in Whole_Sheet:
            cell.font = Font(name='Arial', size=10)
    
    for A1_A100 in Sector_Report.iter_rows(min_row=1, max_col=1, max_row=100):
        for cell3 in A1_A100:
            if (cell3.value=='TOTAL') | (cell3.value=='Sector'):
                for specific_rows in Sector_Report.iter_rows(min_row=cell3.row, max_col=5,max_row=cell3.row):
                    for cell4 in specific_rows:
                        cell4.font = Font(name='Arial', size=10, color="FFFFFF",bold=True)
                        cell4.alignment = Alignment(horizontal='left',wrap_text = False)
                        cell4.fill = PatternFill("solid", fgColor="1F497D")
    
    for column_cells in Sector_Report.iter_rows(min_row=6, min_col=1, max_col=1,max_row=100):
        Sector_Report.column_dimensions[column_cells[0].column].width = 30
                       
    for Whole_Sheet_number in Sector_Report.iter_rows(min_row=8,min_col=2, max_col=3, max_row=10000):
        for cell12 in Whole_Sheet_number:
            cell12.number_format= '#,##0'
            
            
    for Whole_Sheet_number2 in Sector_Report.iter_rows(min_row=8,min_col=4, max_col=5, max_row=10000):
        for cell13 in Whole_Sheet_number2:
            cell13.number_format= '0.00%'
            
    wb.save(File_Name)
    print("Formatting successful the Sector view for "+Logic)


for Logic in ['YTD']:
#for Logic in ['YTD','QTD','MTD']:
    if Logic =='MTD':
        
        First_Day_PY=First_Day_MTD_PY
        Last_Day_PY=Last_Day_MTD_PY
        First_Day_CY=First_Day_MTD_CY
        Last_Day_CY=Last_Day_MTD_CY
#        print (tabulate([["MTD",First_Day_MTD_CY, Last_Day_MTD_CY,First_Day_MTD_PY,Last_Day_MTD_PY]], headers=['MTD/QTD/YTD?','CY SD','CY ED','PY SD','PY ED'], tablefmt='orgtbl'))
        print (tabulate([["MTD",First_Day_PY, Last_Day_PY,First_Day_CY,Last_Day_CY]], headers=['MTD/QTD/YTD?','CY SD','CY ED','PY SD','PY ED'], tablefmt='orgtbl'))
        MFSRGLogic(First_Day_PY,Last_Day_PY,First_Day_CY,Last_Day_CY)
        
    elif Logic =='QTD':
        
        First_Day_PY=First_Day_QTD_PY
        Last_Day_PY=Last_Day_QTD_PY
        First_Day_CY=First_Day_QTD_CY
        Last_Day_CY=Last_Day_QTD_CY
        print (tabulate([["QTD",First_Day_PY, Last_Day_PY,First_Day_CY,Last_Day_CY]], headers=['MTD/QTD/YTD?','CY SD','CY ED','PY SD','PY ED'], tablefmt='orgtbl'))
        MFSRGLogic(First_Day_PY,Last_Day_PY,First_Day_CY,Last_Day_CY)
        
    elif Logic =='YTD':
        
        First_Day_PY=First_Day_YTD_PY
        Last_Day_PY=Last_Day_YTD_PY
        First_Day_CY=First_Day_YTD_CY
        Last_Day_CY=Last_Day_YTD_CY
        print (tabulate([["YTD",First_Day_PY, Last_Day_PY,First_Day_CY,Last_Day_CY]], headers=['MTD/QTD/YTD?','CY SD','CY ED','PY SD','PY ED'], tablefmt='orgtbl'))
        MFSRGLogic(First_Day_PY,Last_Day_PY,First_Day_CY,Last_Day_CY)
        
    else:
        break
